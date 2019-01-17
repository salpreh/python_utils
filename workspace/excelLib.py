# -- encoding: utf-8 --
import openpyxl
import re
from pathlib import Path


def getColumnId(index):
    """
    Return the Id. for the given column index in a spreadsheet
    """
    if index < 1:
        raise ValueError("Column index starts with 1")

    numColumnIndex =  ord('Z') - ord('A') + 1
    columnName = []
    index -= 1
    while index > numColumnIndex - 1:
        columnName.insert(0, chr(ord('A') + index % numColumnIndex))
        index = index // numColumnIndex - 1

    columnName.insert(0, chr(ord('A') + index))

    return ''.join(columnName)

def dataToExcel(excelFilePath, dataList, headerList=[]):
    """
    Escribe los datos de una lista en un fichero excel. Cada lista será una fila y 
    cada dato se colocará en una columna. Opcionalmente se puede proporcionar una
    lista con titulos de cabecera.
    """

    # Comprobaciones en fichero de salida
    excelPath = Path(excelFilePath)

    if not excelPath.parent.exists():
        excelPath.parent.mkdir(parents=True)

    wb = openpyxl.Workbook()
    sheet = wb.active
    columnMaxCell = {}

    # Definición estilos celdas
    borders = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'),
                                     left=openpyxl.styles.Side(style='thin'),
                                     right=openpyxl.styles.Side(style='thin'),
                                     bottom=openpyxl.styles.Side(style='thin'))
    redFill = openpyxl.styles.PatternFill(start_color='f9cfb5',
                                          end_color='f9cfb5',
                                          fill_type='solid')
    headerStyle = openpyxl.styles.NamedStyle('header')
    headerStyle.font = openpyxl.styles.Font(bold=True, size=13)
    headerStyle.border = borders
    headerStyle.fill = openpyxl.styles.PatternFill(start_color='aadcf7',
                                              end_color='aadcf7',
                                              fill_type='solid')

    # Altura de celda
    rowHeight = 20
    
    # Escribimos cabeceras de columnas
    row = 1
    for col, headerTitle in enumerate(headerList, start=1):
        sheet[getColumnId(col)+str(row)] = headerTitle
        sheet[getColumnId(col)+str(row)].style = headerStyle

        # Calculo para tamaño de columnas
        columnMaxCell[getColumnId(col)] = len(headerTitle)

        # Tamaño filas
        sheet.row_dimensions[row].height = rowHeight + 5

    row += 1
    
    # Escribimos resto de líneas
    for dataLine in dataList:
        for col, data in enumerate(dataLine, start=1):
            sheet[getColumnId(col)+str(row)] = data
            sheet[getColumnId(col)+str(row)].border = borders
            sheet[getColumnId(col)+str(row)].fill = redFill

            # Calculo para tamaño de columnas
            colMaxSize = columnMaxCell.setdefault(getColumnId(col), 0)
            if colMaxSize < len(data):
                columnMaxCell[getColumnId(col)] = len(data)
                
        sheet.row_dimensions[row].height = rowHeight
        row += 1

    # Ajustar tamaño de columnas
    for col, size in columnMaxCell.items():
        sheet.column_dimensions[col].width = max(int(size*1.7), 6)
        
    # Guardamos excel
    if not excelPath.name.endswith(".xlsx"):
        excelPath = excelPath.parent / "{}.xlsx".format(excelPath.stem)
    wb.save(excelPath)
    print("> Excel generated in {}".format(excelPath.resolve()))


